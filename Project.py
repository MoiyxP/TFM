import json
import os
from pathlib import Path
from typing import Dict, NamedTuple, Tuple
import numpy as np
import pandas as pd
from pyomeca import Markers, Analogs
from scipy.spatial.transform import Rotation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import detectar_protocolo
 

ruta_excel = r"C:\Users\DanielIordanov\Desktop\VictorMV-biomech\TFM\Code\TFM\Archivos\prueba.xlsx"
static_c3d = r"C:\Users\DanielIordanov\Desktop\VictorMV-biomech\TFM\Code\TFM\Archivos\Tutorial modelos\Visual3D_Tutorial1_SampleC3DFiles\Sample_C3D_Files\Sub01_StandingStaticCal01.c3d"
movement_c3d = r"C:\Users\DanielIordanov\Desktop\VictorMV-biomech\TFM\Code\TFM\Archivos\Tutorial modelos\Visual3D_Tutorial1_SampleC3DFiles\Sample_C3D_Files\Sub01_Walk001.c3d"
CARPETA_SALIDA = "salida_test"

# ----------------------------------
# Lógica del código.
# ----------------------------------

# -------------- Funciones de lectura y calidad


def read_c3d(ruta):
    """
    Lee un archivo .c3d y extrae los marcadores y, si existen, las señales
    analógicas.
    """
    markers = Markers.from_c3d(ruta)
    df_markers = markers.meca.to_wide_dataframe()
    df_markers.to_csv("markers_c3d.csv", sep=';', decimal=',', encoding='utf-8-sig')
 
    try:
        analogs = Analogs.from_c3d(ruta)
    except Exception:
        print("No analogs in this file")
        analogs = None
 
    if analogs is not None:
        df_analogs = analogs.meca.to_wide_dataframe()
        df_analogs.to_csv("analogs_c3d.csv", sep=';', decimal=',', encoding='utf-8-sig')
 
    return markers, analogs
 
def import_excel_data(ruta):
    """
    Lee todas las hojas de un Excel y las organiza en un diccionario.
    Si la hoja tiene 'Tiempo(s)', lo pone como índice.
    """
    dict_hojas = pd.read_excel(ruta, sheet_name=None)
 
    for nombre, df in dict_hojas.items():
        if 'Tiempo(s)' in df.columns:
            df.set_index('Tiempo(s)', inplace=True)
            print(f"Hoja '{nombre}': Detectada como serie temporal.")
        else:
            print(f"Hoja '{nombre}': Detectada como tabla de datos/metadata.")
 
    return dict_hojas

def gap_report(markers, thresholds=(10, 50)):
    """
    Genera un informe de gaps para todos los marcadores de un trial.
 
    """
    marker_names = markers.channel.values.tolist()

    def evaluate_marker_gaps(markers, marker_name, thresholds=(10, 50)):
        """
        Evalúa la calidad de un marcador 3D detectando y clasificando gaps.

        """
    
        small_threshold, medium_threshold = thresholds
    
        # ------ Máscara de NaNs
        
        data = markers.sel(channel=marker_name).values
        mask = np.isnan(data).any(axis=0)  # True donde el frame tiene algún NaN
    
        n_frames = len(mask)
    
        # ------ Pérdida global

        loss_fraction = float(mask.mean())
    
        # ------ Detectar gaps (transiciones)

        changes = np.diff(mask.astype(int))
    
        gap_starts = np.where(changes == 1)[0] + 1
        gap_ends = np.where(changes == -1)[0] + 1
    
        # bordes: gap al inicio o al final del trial
        if mask[0]:
            gap_starts = np.r_[0, gap_starts]
        if mask[-1]:
            gap_ends = np.r_[gap_ends, n_frames]
    
        # ------ Caso sin gaps 

        if len(gap_starts) == 0:
            return {
                "marker": marker_name,
                "loss_fraction": loss_fraction,
                "n_gaps": 0,
                "gap_lengths": [],
                "gap_details": []
            }
    
        # ------ Longitudes de gaps

        gap_lengths = gap_ends - gap_starts
    
        # ------ Análisis por gap

        gap_details = []
    
        for i, (s, e, length) in enumerate(zip(gap_starts, gap_ends, gap_lengths)):
    
            fraction = length / n_frames
    
            if length < small_threshold:
                severity = "small"
            elif length < medium_threshold:
                severity = "medium"
            else:
                severity = "large"
    
            gap_details.append({
                "gap_id": i,
                "start": int(s),
                "end": int(e),
                "length_frames": int(length),
                "fraction_of_trial": float(fraction),
                "severity": severity
            })
    
        # ------ Resultado final
        return {
            "marker": marker_name,
            "loss_fraction": loss_fraction,
            "n_gaps": len(gap_lengths),
            "gap_lengths": gap_lengths.tolist(),
            "gap_details": gap_details
        }
    
    return {
        name: evaluate_marker_gaps(markers, name, thresholds=thresholds)
        for name in marker_names
    }

def interpolate_c3d(markers, analogs, method):
    """
    Vamos a interpolar los datos de los marcadores
    independientemente de la longitud de sus huecos
    """
    int_markers = markers.interpolate_na(dim="time", method=method)
    int_markers = int_markers.ffill(dim="time")
    int_markers = int_markers.bfill(dim="time")
 
    df_markers = int_markers.meca.to_wide_dataframe()
    df_markers.to_csv("int_markers_c3d.csv", sep=';', decimal=',', encoding='utf-8-sig')
 
    if analogs is not None and analogs.size > 0:
        try:
            int_analogs = analogs.interpolate_na(dim="time", method=method)
            int_analogs = int_analogs.ffill(dim="time").bfill(dim="time")
 
            df_analogs = int_analogs.meca.to_wide_dataframe()
            df_analogs.to_csv("int_analogs_c3d.csv", sep=';', decimal=',', encoding='utf-8-sig')
        except Exception as e:
            print(f"Error al procesar analógicos: {e}")
    else:
        print("No se detectaron datos analógicos en este archivo.")
        int_analogs = 'None'
 
    return int_markers, int_analogs

# -------------- Funciones de análisis del c3d 

class SegmentDef(NamedTuple):
    origin: Tuple[str, ...]
    axis_1: Tuple[str, str]
    axis_2: Tuple[str, str]
    axes_name: str
    axis_to_recalculate: str
 
 
# SEGMENTS y XSENS_NAME_MAP se asignan dinámicamente dentro de main(),
# tras detectar el protocolo del c3d. Se inicializan en None acá para que
# las funciones de este módulo (construir_matriz_rotacion, etc.) puedan
# referenciarlas como globales; NO usar estos nombres antes de llamar a
# main() o detectar_protocolo.detectar_y_cargar() explícitamente.
SEGMENTS: Dict[str, SegmentDef] = None
XSENS_NAME_MAP: Dict[str, str] = None

def _vector_entre_marcadores(markers, desde, hasta):
    """Vector (3, n_frames) numpy puro entre dos marcadores, para todos los frames."""
    p_desde = markers.sel(channel=[desde]).values[:3, 0, :]
    p_hasta = markers.sel(channel=[hasta]).values[:3, 0, :]
    return p_hasta - p_desde
 
 
def construir_matriz_rotacion(markers, seg: SegmentDef):
    """
    Construye la matriz de rotación (3, 3, n_frames) de un segmento usando
    el método de los dos vectores (cross product) para garantizar una
    base ortonormal, replicando el algoritmo de pyomeca.Rototrans en
    numpy puro (evita un bug de incompatibilidad de pyomeca con numpy>=2.x).
 
    Lógica del método (importante para que axis_to_recalculate sea
    consistente): los dos vectores MEDIDOS (axis_1, axis_2) casi nunca son
    perfectamente perpendiculares entre sí (son geometría real con ruido).
    El procedimiento correcto es:
        1. Calcular el tercer eje (el que NO viene de axis_1/axis_2) por
           producto cruzado de los dos medidos - ese tercer eje SÍ queda
           perpendicular a ambos, por construcción.
        2. Descartar uno de los DOS EJES MEDIDOS (el menos confiable
           anatómicamente) y reconstruirlo cruzando el tercer eje (ya
           bueno) con el eje medido que se conserva.
    axis_to_recalculate debe ser, por lo tanto, uno de los dos ejes que
    vinieron de axis_1/axis_2 (NO el tercero que ya se calculó por cruce
    en el paso 1) - de lo contrario, el eje recalculado vuelve a salir
    igual al ya calculado, y los otros dos ejes (los medidos) quedan sin
    garantía de ortogonalidad entre sí.
    """
    vector_1 = _vector_entre_marcadores(markers, seg.axis_1[0], seg.axis_1[1])
    vector_2 = _vector_entre_marcadores(markers, seg.axis_2[0], seg.axis_2[1])
    n_frames = vector_1.shape[1]
 
    axes_name = seg.axes_name
    assert axes_name == "".join(sorted(axes_name)), "axes_name debe estar en orden alfabético"
    assert seg.axis_to_recalculate in axes_name, (
        f"axis_to_recalculate='{seg.axis_to_recalculate}' debe ser uno de los ejes "
        f"medidos en axes_name='{axes_name}' (el tercer eje ya se calcula por cruce "
        f"automáticamente, no hace falta pedirlo)"
    )
 
    # Paso 1: asignar los vectores medidos y calcular el tercer eje (bueno)
    if axes_name == "xy":
        x, y = vector_1, vector_2
        z = np.cross(x, y, axis=0)
    elif axes_name == "xz":
        x, z = vector_1, vector_2
        y = np.cross(z, x, axis=0)
    elif axes_name == "yz":
        y, z = vector_1, vector_2
        x = np.cross(y, z, axis=0)
    else:
        raise ValueError(f"axes_name inválido: {axes_name}")
 
    # Paso 2: descartar uno de los DOS EJES MEDIDOS y reconstruirlo con el
    # tercer eje (ya ortogonal) + el eje medido que se conserva.
    ejes = {"x": x, "y": y, "z": z}
    eje_a_descartar = seg.axis_to_recalculate
    eje_conservado = [e for e in axes_name if e != eje_a_descartar][0]
    tercer_eje = [e for e in "xyz" if e not in axes_name][0]
 
    # orden del cruce para mantener la "mano derecha" (x=y×z, y=z×x, z=x×y)
    orden = {"x": ("y", "z"), "y": ("z", "x"), "z": ("x", "y")}
    a, b = orden[eje_a_descartar]
    ejes[eje_a_descartar] = np.cross(ejes[a], ejes[b], axis=0)
 
    x, y, z = ejes["x"], ejes["y"], ejes["z"]
 
    def normalizar(v):
        norm = np.linalg.norm(v, axis=0)
        with np.errstate(invalid="ignore", divide="ignore"):
            return v / norm
 
    x_n, y_n, z_n = normalizar(x), normalizar(y), normalizar(z)
 
    R = np.zeros((3, 3, n_frames))
    R[:, 0, :] = x_n
    R[:, 1, :] = y_n
    R[:, 2, :] = z_n
    return R
 
 
def matriz_a_cuaterniones(R) -> pd.DataFrame:
    """R: (3, 3, n_frames) -> DataFrame QuatW, QuatX, QuatY, QuatZ (scalar-first)."""
    rot_matrices = R.transpose(2, 0, 1)
    valid = ~np.isnan(rot_matrices).any(axis=(1, 2))
 
    quats_wxyz = np.full((rot_matrices.shape[0], 4), np.nan)
    if valid.any():
        r = Rotation.from_matrix(rot_matrices[valid])
        q_xyzw = r.as_quat()
        quats_wxyz[valid] = q_xyzw[:, [3, 0, 1, 2]]
 
    return pd.DataFrame(quats_wxyz, columns=["QuatW", "QuatX", "QuatY", "QuatZ"])
 
 
def matriz_a_euler_zxy(R) -> pd.DataFrame:
    """R: (3, 3, n_frames) -> DataFrame Euler_Z, Euler_X, Euler_Y (grados, secuencia intrínseca ZXY)."""
    rot_matrices = R.transpose(2, 0, 1)
    valid = ~np.isnan(rot_matrices).any(axis=(1, 2))
 
    euler_zxy = np.full((rot_matrices.shape[0], 3), np.nan)
    if valid.any():
        r = Rotation.from_matrix(rot_matrices[valid])
        euler_zxy[valid] = r.as_euler("ZXY", degrees=True)
 
    return pd.DataFrame(euler_zxy, columns=["Euler_Z", "Euler_X", "Euler_Y"])
 
 
def calcular_orientaciones_desde_markers(markers) -> Dict[str, pd.DataFrame]:
    """
    Recibe un objeto Markers ya interpolado y devuelve
    {nombre_segmento: DataFrame con Tiempo(s), cuaterniones y Euler ZXY},
    SIN calibrar contra ningún estático (orientación absoluta "cruda").
    """
    time = markers.time.values
    resultados = {}
    for seg_name, seg_def in SEGMENTS.items():
        R = construir_matriz_rotacion(markers, seg_def)
        df_quat = matriz_a_cuaterniones(R)
        df_euler = matriz_a_euler_zxy(R)
        df = pd.concat([df_quat, df_euler], axis=1)
        df.insert(0, "Tiempo(s)", time)
        resultados[seg_name] = df
    return resultados
 

def _rotacion_promedio_ortonormal(R: np.ndarray) -> np.ndarray:
    """
    Promedia varias matrices de rotación (3, 3, n_frames) en el tiempo y
    devuelve UNA matriz de rotación (3, 3) válida.
 
    El promedio aritmético simple de varias matrices de rotación NO es,
    en general, una rotación válida (su determinante no es exactamente 1).
    Por eso se re-ortonormaliza con SVD: R_ortho = U @ Vt, que es la
    rotación más cercana (en norma de Frobenius) al promedio aritmético.
    Es el procedimiento estándar para promediar orientaciones/posturas
    estáticas en biomecánica.
    """
    R_prom = np.nanmean(R, axis=2)
    U, _, Vt = np.linalg.svd(R_prom)
    R_ortho = U @ Vt
    # Si el SVD da una reflexión (det=-1) en vez de rotación, corregir signo
    if np.linalg.det(R_ortho) < 0:
        U[:, -1] *= -1
        R_ortho = U @ Vt
    return R_ortho
 
 
def calcular_R_estatico_por_segmento(markers_estatico) -> Dict[str, np.ndarray]:
    """
    Calcula, para cada uno de los 7 segmentos, UNA matriz de rotación de
    referencia (3, 3) a partir del trial estático en N-pose: la
    orientación "neutra" de cada segmento según los ejes definidos en
    SEGMENTS, promediada en el tiempo (idealmente el sujeto está quieto).
    """
    R_estatico = {}
    for seg_name, seg_def in SEGMENTS.items():
        R = construir_matriz_rotacion(markers_estatico, seg_def)
        R_estatico[seg_name] = _rotacion_promedio_ortonormal(R)
    return R_estatico
 
 
def calcular_orientaciones_calibradas(
    markers_dinamico,
    R_estatico: Dict[str, np.ndarray],
) -> Dict[str, pd.DataFrame]:
    """
    Calcula cuaterniones y Euler ZXY del trial dinámico, CALIBRADOS contra
    la N-pose estática:
 
        R_calibrado(t) = R_estatico_seg⁻¹ · R_dinamico_seg(t)
 
    Esto resta el offset de construcción geométrica de los ejes locales
    (el que vimos como ángulos absolutos grandes, ej. 40-130°), dejando
    una orientación que parte de ~0° en la postura neutra y se mueve
    desde ahí - comparable con cómo Xsens calibra internamente contra su
    propia N-pose.
 
    Como R_estatico es una matriz de rotación (ortonormal), su inversa es
    su transpuesta (R⁻¹ = Rᵗ), que es lo que se usa aquí.
    """
    time = markers_dinamico.time.values
    resultados = {}
    for seg_name, seg_def in SEGMENTS.items():
        R_din = construir_matriz_rotacion(markers_dinamico, seg_def)  # (3,3,n_frames)
        R_est = R_estatico[seg_name]                                   # (3,3)
        R_est_inv = R_est.T                                             # inversa de una rotación = transpuesta
 
        # R_calibrado(t) = R_est_inv @ R_din(t), para cada frame t
        R_calibrado = np.einsum("ij,jkt->ikt", R_est_inv, R_din)
 
        df_quat = matriz_a_cuaterniones(R_calibrado)
        df_euler = matriz_a_euler_zxy(R_calibrado)
        df = pd.concat([df_quat, df_euler], axis=1)
        df.insert(0, "Tiempo(s)", time)
        resultados[seg_name] = df
    return resultados

# -------------- Funciones de exportación 

def exportar_resultados_excel(resultados: Dict[str, pd.DataFrame], ruta_salida: str):
    """
    Guarda un Excel con una hoja por segmento (PV, TH_R, TH_L, SH_R, SH_L,
    FT_R, FT_L). Cada hoja tiene las columnas Tiempo(s), QuatW, QuatX,
    QuatY, QuatZ, Euler_Z, Euler_X, Euler_Y, cada una en su propia
    columna, una al lado de la otra - nada apretado en una sola celda.
 
    Los números se guardan como valores numéricos reales (no texto), con
    number_format de 6 decimales. El separador decimal que se VE (punto o
    coma) depende de la configuración regional de Excel al abrir el
    archivo - en un Excel configurado en español, se muestra con coma
    automáticamente, sin perder la naturaleza numérica del dato (se puede
    seguir graficando, usando en fórmulas, etc.).
    """
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja vacía por defecto
 
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
 
    for seg_name, df in resultados.items():
        ws = wb.create_sheet(title=seg_name)
 
        # Encabezados: una columna por variable
        for col_idx, col_name in enumerate(df.columns, start=1):
            celda = ws.cell(row=1, column=col_idx, value=col_name)
            celda.font = header_font
            celda.alignment = header_align
 
        # Datos: cada variable mantiene su propia columna, fila por frame
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, valor in enumerate(row, start=1):
                celda = ws.cell(row=row_idx, column=col_idx, value=float(valor))
                celda.number_format = "0.000000"
 
        # Ancho de columna ajustado al contenido (encabezado como referencia)
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(col_name) + 2)
 
        ws.freeze_panes = "A2"  # encabezado siempre visible al hacer scroll
 
    wb.save(ruta_salida)

# ----------------------------------
# Aplicación del código.
# ----------------------------------

def main():
    carpeta_salida = Path(CARPETA_SALIDA)
    carpeta_salida.mkdir(parents=True, exist_ok=True)
 
    cwd_anterior = os.getcwd()
    os.chdir(carpeta_salida)  # read_c3d / interpolate_c3d guardan sus CSV en el cwd
    try:
        # --- 1. Evaluación de calidad del trial de MOVIMIENTO ---
        markers, analogs = read_c3d(movement_c3d)
        os.replace("markers_c3d.csv", "markers_c3d_movimiento.csv")
        if os.path.exists("analogs_c3d.csv"):
            os.replace("analogs_c3d.csv", "analogs_c3d_movimiento.csv")
 
        # Detección automática del protocolo de marcadores (ver
        # detectar_protocolo.py): actualiza las globales SEGMENTS y
        # XSENS_NAME_MAP de este módulo para que el resto del pipeline
        # (construir_matriz_rotacion, calibración, etc.) las use sin
        # necesidad de saber qué protocolo es.
        global SEGMENTS, XSENS_NAME_MAP
        SEGMENTS, XSENS_NAME_MAP = detectar_protocolo.detectar_y_cargar(markers.channel.values)
 
        report = gap_report(markers)
        with open("results.json", "w") as f:  # lo visualizamos en un json
            json.dump(report, f, indent=4)
 
        # --- 2. Interpolación del trial de MOVIMIENTO con 'akima' ---
        # int_markers (interpolado, SIN gaps) es lo que se usa más abajo para
        # calcular las orientaciones - no markers (el crudo, con gaps).
        int_markers, int_analogs = interpolate_c3d(markers, analogs, 'akima')
        os.replace("int_markers_c3d.csv", "int_markers_c3d_movimiento.csv")
        if os.path.exists("int_analogs_c3d.csv"):
            os.replace("int_analogs_c3d.csv", "int_analogs_c3d_movimiento.csv")
 
        # --- 3. Lectura + calidad + interpolación del trial ESTÁTICO (N-pose) ---
        # MISMO procedimiento que el dinámico (pasos 1-2), repetido aquí para
        # el estático, que se usa solo como referencia de calibración (no
        # como dato de movimiento). Los CSV del dinámico YA se renombraron
        # arriba antes de llegar aquí - read_c3d/interpolate_c3d siempre
        # escriben los mismos nombres de archivo (markers_c3d.csv, etc.), así
        # que si no se renombran de inmediato, el estático los pisaría.
        markers_static, analogs_static = read_c3d(static_c3d)
        os.replace("markers_c3d.csv", "markers_c3d_static.csv")
        if os.path.exists("analogs_c3d.csv"):
            os.replace("analogs_c3d.csv", "analogs_c3d_static.csv")
 
        report_static = gap_report(markers_static)
        with open("results_static.json", "w") as f:
            json.dump(report_static, f, indent=4)
 
        int_markers_static, _ = interpolate_c3d(markers_static, analogs_static, 'akima')
        os.replace("int_markers_c3d.csv", "int_markers_c3d_static.csv")
        if os.path.exists("int_analogs_c3d.csv"):
            os.replace("int_analogs_c3d.csv", "int_analogs_c3d_static.csv")
 
        # --- 4. Importación de los datos del sensor inercial (Excel) ---
        excel_data = import_excel_data(ruta_excel)
 
    finally:
        os.chdir(cwd_anterior)
 
    # --- 5. Orientaciones SIN calibrar (absolutas) del trial de MOVIMIENTO ---
    resultados_crudos = calcular_orientaciones_desde_markers(int_markers)
 
    # --- 5b. Orientaciones del trial ESTÁTICO (para poder revisar la N-pose:
    # confirmar que el sujeto estuvo realmente quieto y simétrico antes de
    # confiar en R_estatico, que se calcula a partir de estas mismas orientaciones) ---
    resultados_estatico = calcular_orientaciones_desde_markers(int_markers_static)
 
    # --- 6. Referencia de calibración: R por segmento desde la N-pose ---
    R_estatico = calcular_R_estatico_por_segmento(int_markers_static)
 
    # --- 7. Orientaciones CALIBRADAS contra la N-pose ---
    resultados_calibrados = calcular_orientaciones_calibradas(int_markers, R_estatico)
 
    # --- 8. Guardado de resultados: tres Excel (crudo, calibrado, estático), ---
    # cada uno con una hoja por segmento y cada variable en su propia columna,
    # todos en la misma carpeta_salida junto con los CSV de marcadores y JSON
    ruta_excel_crudo = carpeta_salida / "qualisys_crudo.xlsx"
    ruta_excel_calibrado = carpeta_salida / "qualisys_calibrado.xlsx"
    ruta_excel_estatico = carpeta_salida / "qualisys_estatico.xlsx"
 
    exportar_resultados_excel(resultados_crudos, ruta_excel_crudo)
    exportar_resultados_excel(resultados_calibrados, ruta_excel_calibrado)
    exportar_resultados_excel(resultados_estatico, ruta_excel_estatico)
 
    print("\n=== Orientaciones por segmento (movimiento crudo, calibrado, y estático) ===")
    print(f"  Crudo     -> {ruta_excel_crudo.name}  (hojas: {', '.join(resultados_crudos.keys())})")
    print(f"  Calibrado -> {ruta_excel_calibrado.name}  (hojas: {', '.join(resultados_calibrados.keys())})")
    print(f"  Estático  -> {ruta_excel_estatico.name}  (hojas: {', '.join(resultados_estatico.keys())})")
 
    return {
        "excel_data": excel_data,
        "resultados_crudos": resultados_crudos,
        "resultados_calibrados": resultados_calibrados,
        "resultados_estatico": resultados_estatico,
        "R_estatico": R_estatico,
    }
 
 
if __name__ == "__main__":
    main()